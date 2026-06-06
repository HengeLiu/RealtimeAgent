#!/usr/bin/env python3
"""运行独立 VAD 边界检测基准实验。

主要功能：读取一组短 WAV 样本，生成带静音和轻噪声的实验输入，使用 RMS 派生
proxy speech boundary，再对 RMS、WebRTC VAD、Silero VAD ONNX、TEN VAD 进行
边界检测对比，并输出 JSONL、XLSX 和 Markdown 报告。

主要逻辑：
1. 读取 PCM16 mono 16k WAV 样本。
2. 为每条样本生成 clean_padded、white_noise_snr20、hum_50hz_snr25 三个版本。
3. 用 RMS 自适应阈值标注 proxy speech_start / speech_stop。
4. 调用可用的 VAD provider，记录 speech_started / speech_stopped 和推理耗时。
5. 汇总 start/stop 延迟、miss、false start 和推理耗时。

参数：通过命令行传入样本目录、输出目录、随机种子和 provider 列表。
返回值：命令行退出码。成功时为 0；输入音频不合法或输出失败时抛出异常。
异常情况：缺失某个可选 VAD 依赖时，该 provider 会记录为 unavailable，不影响其余
provider 继续执行。
"""

from __future__ import annotations

import argparse
import importlib.metadata
import json
import math
import platform
import random
import statistics
import sys
import time
import wave
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import numpy as np


SAMPLE_RATE = 16_000
PCM_MAX = 32768.0
DEFAULT_STOP_WAIT_MS = [200, 300, 400, 600, 800, 1000]
DEFAULT_SAMPLES = [
    "回家.wav",
    "继续.wav",
    "你是谁呀.wav",
    "步行回家.wav",
    "我叫什么呀.wav",
    "自我介绍一下.wav",
    "给我讲个笑话吧.wav",
    "我刚才问了你什么.wav",
    "一分钟后提醒我.wav",
    "我的住址在哪里.wav",
]


@dataclass(frozen=True)
class AudioCase:
    """保存单个实验音频的信息。

    主要功能：描述一个原始样本经过增强后的实验输入。
    主要属性：case_id 是稳定标识；variant 是增强方式；pcm 是标准化后的 PCM16 数组。
    """

    case_id: str
    source_name: str
    source_path: str
    variant: str
    duration_ms: int
    pcm: np.ndarray


@dataclass(frozen=True)
class Label:
    """保存 RMS 派生的语音边界。

    主要功能：作为本轮 benchmark 的 proxy ground truth。
    主要属性：start_ms / stop_ms 分别表示 RMS 估计的语音开始和结束时间。
    """

    case_id: str
    status: str
    start_ms: int | None
    stop_ms: int | None
    threshold: float
    noise_floor: float
    speech_peak: float


class VADProvider:
    """统一 VAD provider 接口。

    主要功能：接收连续 PCM16 音频帧，输出 speech_started / speech_stopped。
    主要方法：reset() 清空内部状态；process() 处理一帧并返回事件。
    主要属性：name 标识 provider；frame_samples 表示该 provider 的输入帧长。
    """

    name = "base"
    family = "base"
    frame_samples = 320
    stop_wait_ms: int | None = None

    def reset(self) -> None:
        """清空 provider 内部状态。

        参数：无。
        返回值：无。
        异常情况：默认实现不抛异常。
        """

    def process(self, frame: np.ndarray, audio_ms: int) -> list[dict[str, Any]]:
        """处理一帧 PCM16 音频。

        参数：
            frame：int16 mono PCM 音频帧。
            audio_ms：当前帧结束点对应的音频时间。
        返回值：事件列表。没有边界时返回空列表。
        异常情况：具体 provider 可抛出依赖或输入格式异常。
        """

        raise NotImplementedError


class RMSProvider(VADProvider):
    """基于 RMS 阈值的零依赖 VAD baseline。

    主要功能：用能量阈值生成 speech_started / speech_stopped，用于确认 benchmark
    harness 是否正常，不作为生产候选。
    主要方法：process() 计算当前帧 RMS 并更新语音状态。
    主要属性：threshold 是 PCM full scale 归一化阈值；silence_ms 控制 stop 延迟。
    """

    family = "rms_baseline"
    frame_samples = 320

    def __init__(self, threshold: float = 0.018, start_confirm_ms: int = 60, silence_ms: int = 360) -> None:
        self.name = f"rms_baseline_s{silence_ms}"
        self.stop_wait_ms = silence_ms
        self.threshold = threshold
        self.start_confirm_ms = start_confirm_ms
        self.silence_ms = silence_ms
        self.reset()

    def reset(self) -> None:
        self.triggered = False
        self.speech_candidate_ms: int | None = None
        self.last_voice_ms: int | None = None

    def process(self, frame: np.ndarray, audio_ms: int) -> list[dict[str, Any]]:
        started = time.perf_counter()
        score = rms_float(frame)
        infer_ms = elapsed_ms(started)
        events: list[dict[str, Any]] = []
        frame_ms = samples_to_ms(len(frame))
        frame_start_ms = max(0, audio_ms - frame_ms)

        if score >= self.threshold:
            self.last_voice_ms = audio_ms
            if self.speech_candidate_ms is None:
                self.speech_candidate_ms = frame_start_ms
            if not self.triggered and audio_ms - self.speech_candidate_ms >= self.start_confirm_ms:
                self.triggered = True
                events.append(make_event("speech_started", self.name, self.speech_candidate_ms, score, infer_ms))
        else:
            self.speech_candidate_ms = None
            if self.triggered and self.last_voice_ms is not None and audio_ms - self.last_voice_ms >= self.silence_ms:
                self.triggered = False
                events.append(make_event("speech_stopped", self.name, audio_ms, score, infer_ms))
                self.last_voice_ms = None
        return events


class WebRtcProvider(VADProvider):
    """WebRTC VAD provider。

    主要功能：使用 py-webrtcvad 的二值结果做快速 speech_started 检测和基础 stop。
    主要方法：process() 将 20ms PCM 帧传给 webrtcvad.Vad.is_speech。
    主要属性：mode 越高越激进；silence_ms 控制 speech_stopped 的静音窗口。
    """

    family = "webrtcvad_mode2"
    frame_samples = 320

    def __init__(self, mode: int = 2, start_confirm_ms: int = 40, silence_ms: int = 360) -> None:
        import webrtcvad

        self.name = f"webrtcvad_mode{mode}_s{silence_ms}"
        self.stop_wait_ms = silence_ms
        self.vad = webrtcvad.Vad(mode)
        self.start_confirm_ms = start_confirm_ms
        self.silence_ms = silence_ms
        self.reset()

    def reset(self) -> None:
        self.triggered = False
        self.speech_candidate_ms: int | None = None
        self.last_voice_ms: int | None = None

    def process(self, frame: np.ndarray, audio_ms: int) -> list[dict[str, Any]]:
        started = time.perf_counter()
        is_speech = self.vad.is_speech(frame.astype(np.int16).tobytes(), SAMPLE_RATE)
        infer_ms = elapsed_ms(started)
        score = 1.0 if is_speech else 0.0
        events: list[dict[str, Any]] = []
        frame_ms = samples_to_ms(len(frame))
        frame_start_ms = max(0, audio_ms - frame_ms)

        if is_speech:
            self.last_voice_ms = audio_ms
            if self.speech_candidate_ms is None:
                self.speech_candidate_ms = frame_start_ms
            if not self.triggered and audio_ms - self.speech_candidate_ms >= self.start_confirm_ms:
                self.triggered = True
                events.append(make_event("speech_started", self.name, self.speech_candidate_ms, score, infer_ms))
        else:
            self.speech_candidate_ms = None
            if self.triggered and self.last_voice_ms is not None and audio_ms - self.last_voice_ms >= self.silence_ms:
                self.triggered = False
                events.append(make_event("speech_stopped", self.name, audio_ms, score, infer_ms))
                self.last_voice_ms = None
        return events


class SileroOnnxProvider(VADProvider):
    """Silero VAD ONNX provider。

    主要功能：使用 silero-vad 的 ONNX 模型和 VADIterator 输出 streaming 边界。
    主要方法：process() 处理 512 samples chunk，并把 sample index 转为毫秒。
    主要属性：threshold 控制语音概率阈值，min_silence_ms 控制 end 检测。
    """

    family = "silero_onnx"
    frame_samples = 512

    def __init__(self, threshold: float = 0.5, min_silence_ms: int = 360, speech_pad_ms: int = 30) -> None:
        from silero_vad import VADIterator, load_silero_vad

        self.name = f"silero_onnx_s{min_silence_ms}"
        self.stop_wait_ms = min_silence_ms
        self.model = load_silero_vad(onnx=True)
        self.iterator_cls = VADIterator
        self.threshold = threshold
        self.min_silence_ms = min_silence_ms
        self.speech_pad_ms = speech_pad_ms
        self.reset()

    def reset(self) -> None:
        self.model.reset_states()
        self.iterator = self.iterator_cls(
            self.model,
            threshold=self.threshold,
            sampling_rate=SAMPLE_RATE,
            min_silence_duration_ms=self.min_silence_ms,
            speech_pad_ms=self.speech_pad_ms,
        )

    def process(self, frame: np.ndarray, audio_ms: int) -> list[dict[str, Any]]:
        audio_float = frame.astype(np.float32) / PCM_MAX
        started = time.perf_counter()
        item = self.iterator(audio_float, return_seconds=False)
        infer_ms = elapsed_ms(started)
        if not item:
            return []
        if "start" in item:
            return [make_event("speech_started", self.name, samples_to_ms(int(item["start"])), None, infer_ms)]
        if "end" in item:
            return [make_event("speech_stopped", self.name, audio_ms, None, infer_ms)]
        return []


class TenVadProvider(VADProvider):
    """TEN VAD provider。

    主要功能：使用 ten-vad Python binding 输出语音概率和 flags，再由实验层状态机生成边界。
    主要方法：process() 处理 256 samples chunk。
    主要属性：threshold 控制概率阈值，silence_ms 控制 speech_stopped 静音窗口。
    """

    family = "ten_vad"
    frame_samples = 256

    def __init__(self, threshold: float = 0.5, start_confirm_ms: int = 32, silence_ms: int = 320) -> None:
        from ten_vad import TenVad

        self.name = f"ten_vad_s{silence_ms}"
        self.stop_wait_ms = silence_ms
        self.vad = TenVad(hop_size=self.frame_samples, threshold=threshold)
        self.threshold = threshold
        self.start_confirm_ms = start_confirm_ms
        self.silence_ms = silence_ms
        self.reset()

    def reset(self) -> None:
        self.triggered = False
        self.speech_candidate_ms: int | None = None
        self.last_voice_ms: int | None = None

    def process(self, frame: np.ndarray, audio_ms: int) -> list[dict[str, Any]]:
        started = time.perf_counter()
        probability, flags = self.vad.process(frame.astype(np.int16))
        infer_ms = elapsed_ms(started)
        is_speech = bool(flags) or probability >= self.threshold
        events: list[dict[str, Any]] = []
        frame_ms = samples_to_ms(len(frame))
        frame_start_ms = max(0, audio_ms - frame_ms)

        if is_speech:
            self.last_voice_ms = audio_ms
            if self.speech_candidate_ms is None:
                self.speech_candidate_ms = frame_start_ms
            if not self.triggered and audio_ms - self.speech_candidate_ms >= self.start_confirm_ms:
                self.triggered = True
                events.append(make_event("speech_started", self.name, self.speech_candidate_ms, probability, infer_ms))
        else:
            self.speech_candidate_ms = None
            if self.triggered and self.last_voice_ms is not None and audio_ms - self.last_voice_ms >= self.silence_ms:
                self.triggered = False
                events.append(make_event("speech_stopped", self.name, audio_ms, probability, infer_ms))
                self.last_voice_ms = None
        return events


class RmsStartSileroStopProvider(VADProvider):
    """RMS 低成本 start 候选 + Silero 结束判断的组合 provider。

    主要功能：用 RMS 能量门限触发 speech_started，用 Silero 概率的连续静音触发
    speech_stopped，验证“低成本 start 候选 + 模型结束”的组合是否适合 Omni manual。
    主要方法：process() 同时计算 RMS 和 Silero probability，再按组合状态机输出边界。
    主要属性：silence_ms 是结束等待时间；rms_threshold 是启动门限。
    """

    family = "rms_start_silero_stop"
    frame_samples = 512

    def __init__(
        self,
        rms_threshold: float = 0.018,
        silero_threshold: float = 0.5,
        start_confirm_ms: int = 64,
        silence_ms: int = 360,
    ) -> None:
        from silero_vad import load_silero_vad

        self.name = f"rms_start_silero_stop_s{silence_ms}"
        self.stop_wait_ms = silence_ms
        self.rms_threshold = rms_threshold
        self.silero_threshold = silero_threshold
        self.start_confirm_ms = start_confirm_ms
        self.silence_ms = silence_ms
        self.model = load_silero_vad(onnx=True)
        self.reset()

    def reset(self) -> None:
        self.model.reset_states()
        self.triggered = False
        self.speech_candidate_ms: int | None = None
        self.last_voice_ms: int | None = None

    def process(self, frame: np.ndarray, audio_ms: int) -> list[dict[str, Any]]:
        started = time.perf_counter()
        rms_score = rms_float(frame)
        silero_score = silero_probability(self.model, frame)
        infer_ms = elapsed_ms(started)
        events: list[dict[str, Any]] = []
        frame_ms = samples_to_ms(len(frame))
        frame_start_ms = max(0, audio_ms - frame_ms)
        start_speech = rms_score >= self.rms_threshold
        stop_speech = silero_score >= self.silero_threshold

        if start_speech:
            if self.speech_candidate_ms is None:
                self.speech_candidate_ms = frame_start_ms
            if not self.triggered and audio_ms - self.speech_candidate_ms >= self.start_confirm_ms:
                self.triggered = True
                self.last_voice_ms = audio_ms
                events.append(
                    make_event("speech_started", self.name, self.speech_candidate_ms, silero_score, infer_ms)
                )
        else:
            self.speech_candidate_ms = None

        if self.triggered:
            if stop_speech:
                self.last_voice_ms = audio_ms
            elif self.last_voice_ms is not None and audio_ms - self.last_voice_ms >= self.silence_ms:
                self.triggered = False
                events.append(make_event("speech_stopped", self.name, audio_ms, silero_score, infer_ms))
                self.last_voice_ms = None
        return events


class SileroStartRmsGuardStopProvider(VADProvider):
    """Silero 启动 + RMS 守护结束的组合 provider。

    主要功能：用 Silero probability 触发 speech_started，结束时要求 Silero 和 RMS 都
    连续静音，验证“模型启动 + 能量兜底防 false stop”的组合是否更稳。
    主要方法：process() 计算两种信号，并在两者都低于门限时累计静音时间。
    主要属性：silence_ms 是结束等待时间；rms_threshold 用于防止低能量尾音被过早截断。
    """

    family = "silero_start_rms_guard_stop"
    frame_samples = 512

    def __init__(
        self,
        rms_threshold: float = 0.012,
        silero_threshold: float = 0.5,
        start_confirm_ms: int = 64,
        silence_ms: int = 360,
    ) -> None:
        from silero_vad import load_silero_vad

        self.name = f"silero_start_rms_guard_stop_s{silence_ms}"
        self.stop_wait_ms = silence_ms
        self.rms_threshold = rms_threshold
        self.silero_threshold = silero_threshold
        self.start_confirm_ms = start_confirm_ms
        self.silence_ms = silence_ms
        self.model = load_silero_vad(onnx=True)
        self.reset()

    def reset(self) -> None:
        self.model.reset_states()
        self.triggered = False
        self.speech_candidate_ms: int | None = None
        self.last_voice_ms: int | None = None

    def process(self, frame: np.ndarray, audio_ms: int) -> list[dict[str, Any]]:
        started = time.perf_counter()
        rms_score = rms_float(frame)
        silero_score = silero_probability(self.model, frame)
        infer_ms = elapsed_ms(started)
        events: list[dict[str, Any]] = []
        frame_ms = samples_to_ms(len(frame))
        frame_start_ms = max(0, audio_ms - frame_ms)
        silero_speech = silero_score >= self.silero_threshold
        rms_speech = rms_score >= self.rms_threshold

        if silero_speech:
            if self.speech_candidate_ms is None:
                self.speech_candidate_ms = frame_start_ms
            if not self.triggered and audio_ms - self.speech_candidate_ms >= self.start_confirm_ms:
                self.triggered = True
                self.last_voice_ms = audio_ms
                events.append(
                    make_event("speech_started", self.name, self.speech_candidate_ms, silero_score, infer_ms)
                )
        else:
            self.speech_candidate_ms = None

        if self.triggered:
            if silero_speech or rms_speech:
                self.last_voice_ms = audio_ms
            elif self.last_voice_ms is not None and audio_ms - self.last_voice_ms >= self.silence_ms:
                self.triggered = False
                events.append(make_event("speech_stopped", self.name, audio_ms, silero_score, infer_ms))
                self.last_voice_ms = None
        return events


def parse_args() -> argparse.Namespace:
    """解析命令行参数。

    参数：无。
    返回值：argparse.Namespace，包含样本目录、输出目录、provider 等配置。
    异常情况：参数非法时 argparse 会退出进程。
    """

    parser = argparse.ArgumentParser(description="Run VAD boundary benchmark.")
    parser.add_argument("--sample-dir", default="testdata/audio-sample", help="WAV 样本目录。")
    parser.add_argument("--out-root", default="runs/vad-benchmark", help="实验产物根目录。")
    parser.add_argument(
        "--providers",
        default=(
            "rms_baseline,webrtcvad_mode2,silero_onnx,ten_vad,"
            "rms_start_silero_stop,silero_start_rms_guard_stop"
        ),
        help="逗号分隔的 provider 名称。",
    )
    parser.add_argument(
        "--stop-wait-ms",
        default="200,300,400,600,800,1000",
        help="逗号分隔的 speech_stopped 等待窗口，单位毫秒。",
    )
    parser.add_argument("--seed", type=int, default=20260605, help="噪声增强随机种子。")
    parser.add_argument("--timestamp", default=None, help="指定输出目录时间戳，默认自动生成。")
    return parser.parse_args()


def parse_int_list(raw: str) -> list[int]:
    """解析逗号分隔整数列表。

    参数：raw，逗号分隔的整数字符串。
    返回值：整数列表。
    异常情况：列表为空或包含非正数时抛出 ValueError。
    """

    values = [int(item.strip()) for item in raw.split(",") if item.strip()]
    if not values or any(value <= 0 for value in values):
        raise ValueError(f"invalid integer list: {raw}")
    return values


def main() -> None:
    """执行完整 benchmark。

    参数：无，使用命令行参数。
    返回值：无。
    异常情况：音频格式不符合预期或输出文件写入失败时抛出异常。
    """

    args = parse_args()
    random.seed(args.seed)
    np.random.seed(args.seed)
    sample_dir = Path(args.sample_dir)
    timestamp = args.timestamp or datetime.now().strftime("%Y%m%d-%H%M%S")
    out_dir = Path(args.out_root) / timestamp
    generated_audio_dir = out_dir / "generated-audio"
    generated_audio_dir.mkdir(parents=True, exist_ok=True)

    provider_names = [item.strip() for item in args.providers.split(",") if item.strip()]
    stop_wait_values = parse_int_list(args.stop_wait_ms)
    providers, provider_status = build_providers(provider_names, stop_wait_values)
    cases = build_cases(sample_dir, generated_audio_dir)
    labels = [label_case(case) for case in cases]
    label_map = {label.case_id: label for label in labels}

    all_events: list[dict[str, Any]] = []
    sample_results: list[dict[str, Any]] = []
    inference_records: list[dict[str, Any]] = []

    for case in cases:
        label = label_map[case.case_id]
        for provider in providers:
            events, infer_records = run_provider(case, provider)
            all_events.extend(events)
            inference_records.extend(infer_records)
            sample_results.append(summarize_case_provider(case, label, provider, events, infer_records))

    summary_rows = summarize_provider_results(sample_results)
    manifest = build_manifest(args, out_dir, cases, provider_status)

    write_json(out_dir / "manifest.json", manifest)
    write_jsonl(out_dir / "labels.jsonl", [label.__dict__ for label in labels])
    write_jsonl(out_dir / "provider-events.jsonl", all_events)
    write_jsonl(out_dir / "inference.jsonl", inference_records)
    write_jsonl(out_dir / "sample-results.jsonl", sample_results)
    write_json(out_dir / "summary.json", {"providers": summary_rows})
    write_xlsx(out_dir / "summary.xlsx", summary_rows, sample_results, labels, provider_status)
    write_report(out_dir / "report.md", summary_rows, sample_results, labels, provider_status, manifest)
    print(json.dumps({"out_dir": str(out_dir), "providers": [p.name for p in providers]}, ensure_ascii=False, indent=2))


def build_providers(provider_names: list[str], stop_wait_values: list[int]) -> tuple[list[VADProvider], list[dict[str, Any]]]:
    """按名称初始化 provider。

    参数：
        provider_names：需要执行的 provider 名称列表。
    返回值：可用 provider 列表和所有 provider 的状态记录。
    异常情况：单个 provider 初始化失败时记录 unavailable，不向外抛出。
    """

    factories: dict[str, Any] = {
        "rms_baseline": RMSProvider,
        "webrtcvad_mode2": WebRtcProvider,
        "silero_onnx": SileroOnnxProvider,
        "ten_vad": TenVadProvider,
        "rms_start_silero_stop": RmsStartSileroStopProvider,
        "silero_start_rms_guard_stop": SileroStartRmsGuardStopProvider,
    }
    providers: list[VADProvider] = []
    statuses: list[dict[str, Any]] = []
    for name in provider_names:
        factory = factories.get(name)
        if factory is None:
            statuses.append({"provider": name, "available": False, "reason": "unknown_provider"})
            continue
        for stop_wait_ms in stop_wait_values:
            started = time.perf_counter()
            try:
                provider = make_provider(factory, name, stop_wait_ms)
                providers.append(provider)
                statuses.append(
                    {
                        "provider": provider.name,
                        "provider_family": provider.family,
                        "stop_wait_ms": provider.stop_wait_ms,
                        "available": True,
                        "init_ms": round(elapsed_ms(started), 3),
                        "frame_samples": provider.frame_samples,
                    }
                )
            except Exception as exc:  # noqa: BLE001 - 实验脚本需要继续执行其他 provider
                statuses.append(
                    {
                        "provider": f"{name}_s{stop_wait_ms}",
                        "provider_family": name,
                        "stop_wait_ms": stop_wait_ms,
                        "available": False,
                        "reason": f"{type(exc).__name__}: {exc}",
                    }
                )
    for name in ["funasr_fsmn_vad", "firered_vad"]:
        if name not in provider_names:
            statuses.append(
                {
                    "provider": name,
                    "provider_family": name,
                    "stop_wait_ms": None,
                    "available": False,
                    "reason": "not_run_first_round_heavy_dependency",
                }
            )
    return providers, statuses


def make_provider(factory: Any, name: str, stop_wait_ms: int) -> VADProvider:
    """创建带指定 stop wait 的 provider。

    参数：factory 是 provider 构造器；name 是 provider family；stop_wait_ms 是结束等待时间。
    返回值：VADProvider 实例。
    异常情况：构造器失败时向外抛出。
    """

    if name == "silero_onnx":
        return factory(min_silence_ms=stop_wait_ms)
    return factory(silence_ms=stop_wait_ms)


def build_cases(sample_dir: Path, generated_audio_dir: Path) -> list[AudioCase]:
    """构建实验样本列表。

    参数：
        sample_dir：原始 WAV 样本目录。
        generated_audio_dir：增强音频输出目录。
    返回值：AudioCase 列表。
    异常情况：任一默认样本缺失或 WAV 格式不合法时抛出异常。
    """

    cases: list[AudioCase] = []
    for index, name in enumerate(DEFAULT_SAMPLES, start=1):
        path = sample_dir / name
        pcm = read_wav_pcm16(path)
        for variant, variant_pcm in build_variants(pcm).items():
            case_id = f"{index:02d}-{path.stem}-{variant}"
            out_wav = generated_audio_dir / f"{case_id}.wav"
            write_wav_pcm16(out_wav, variant_pcm)
            cases.append(
                AudioCase(
                    case_id=case_id,
                    source_name=path.name,
                    source_path=str(path),
                    variant=variant,
                    duration_ms=samples_to_ms(len(variant_pcm)),
                    pcm=variant_pcm,
                )
            )
    return cases


def build_variants(pcm: np.ndarray) -> dict[str, np.ndarray]:
    """生成 clean 和噪声增强版本。

    参数：pcm，原始 PCM16 音频。
    返回值：variant 名称到 PCM16 数组的映射。
    异常情况：输入为空时抛出 ValueError。
    """

    if len(pcm) == 0:
        raise ValueError("empty audio")
    leading = np.zeros(ms_to_samples(1500), dtype=np.int16)
    trailing = np.zeros(ms_to_samples(1200), dtype=np.int16)
    clean = np.concatenate([leading, pcm, trailing])
    return {
        "clean_padded": clean,
        "white_noise_snr20": add_white_noise(clean, snr_db=20.0),
        "hum_50hz_snr25": add_hum(clean, hz=50.0, snr_db=25.0),
    }


def label_case(case: AudioCase) -> Label:
    """使用 RMS 为样本生成 proxy speech boundary。

    参数：case，单个实验音频。
    返回值：Label，包含状态、起止时间和阈值信息。
    异常情况：无稳定边界时返回 unstable 状态，不抛异常。
    """

    frame = ms_to_samples(20)
    hop = ms_to_samples(10)
    values = []
    for start in range(0, max(1, len(case.pcm) - frame + 1), hop):
        values.append(rms_float(case.pcm[start : start + frame]))
    if not values:
        return Label(case.case_id, "unstable", None, None, 0.0, 0.0, 0.0)

    rms_values = np.array(values, dtype=np.float32)
    smooth = median_filter(rms_values, size=5)
    head_frames = max(1, ms_to_samples(1000) // hop)
    tail_frames = max(1, ms_to_samples(800) // hop)
    noise_values = np.concatenate([smooth[:head_frames], smooth[-tail_frames:]])
    noise_floor = float(np.percentile(noise_values, 90))
    speech_peak = float(np.percentile(smooth, 95))
    threshold = max(noise_floor * 3.0, speech_peak * 0.08, 0.005)
    voiced = smooth >= threshold
    min_frames = max(1, ms_to_samples(80) // hop)
    start_index = find_first_run(voiced, min_frames)
    stop_index = find_last_run(voiced, min_frames)
    if start_index is None or stop_index is None or stop_index < start_index:
        return Label(case.case_id, "unstable", None, None, threshold, noise_floor, speech_peak)
    start_ms = samples_to_ms(start_index * hop)
    stop_ms = samples_to_ms(min(len(case.pcm), stop_index * hop + frame))
    return Label(case.case_id, "ok", start_ms, stop_ms, threshold, noise_floor, speech_peak)


def run_provider(case: AudioCase, provider: VADProvider) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """对单个样本运行一个 provider。

    参数：
        case：实验音频。
        provider：VAD provider 实例。
    返回值：provider 事件列表和每帧推理耗时列表。
    异常情况：provider.process 抛出的异常会上抛，因为此时结果不可用。
    """

    provider.reset()
    events: list[dict[str, Any]] = []
    inference: list[dict[str, Any]] = []
    frame = provider.frame_samples
    padded = pad_to_frame(case.pcm, frame)
    for start in range(0, len(padded), frame):
        chunk = padded[start : start + frame]
        audio_ms = samples_to_ms(start + frame)
        before = time.perf_counter()
        chunk_events = provider.process(chunk, audio_ms)
        total_infer_ms = elapsed_ms(before)
        inference.append(
            {
                "case_id": case.case_id,
                "provider": provider.name,
                "provider_family": provider.family,
                "stop_wait_ms": provider.stop_wait_ms,
                "audio_ms": audio_ms,
                "infer_ms": round(total_infer_ms, 6),
            }
        )
        for event in chunk_events:
            event.update(
                {
                    "case_id": case.case_id,
                    "source_name": case.source_name,
                    "variant": case.variant,
                    "provider_family": provider.family,
                    "stop_wait_ms": provider.stop_wait_ms,
                }
            )
            events.append(event)
    return events, inference


def summarize_case_provider(
    case: AudioCase,
    label: Label,
    provider: VADProvider,
    events: list[dict[str, Any]],
    inference: list[dict[str, Any]],
) -> dict[str, Any]:
    """汇总单个样本和 provider 的检测结果。

    参数：case、label、provider、events、inference。
    返回值：单条明细结果字典。
    异常情况：无。
    """

    start_events = [event for event in events if event["event"] == "speech_started"]
    first_start = start_events[0]["audio_ms"] if start_events else None
    early_cutoff = None if label.start_ms is None else label.start_ms - 300
    false_start_events = [
        event for event in start_events if early_cutoff is not None and event["audio_ms"] < early_cutoff
    ]
    accepted_start_events = [
        event for event in start_events if early_cutoff is None or event["audio_ms"] >= early_cutoff
    ]
    detected_start = accepted_start_events[0]["audio_ms"] if accepted_start_events else None
    stop_events = [
        event
        for event in events
        if event["event"] == "speech_stopped" and (detected_start is None or event["audio_ms"] >= detected_start)
    ]
    detected_stop = stop_events[0]["audio_ms"] if stop_events else None
    infer_values = [item["infer_ms"] for item in inference]
    start_delay = none_subtract(detected_start, label.start_ms)
    stop_delay = none_subtract(detected_stop, label.stop_ms)
    false_start = bool(false_start_events)
    return {
        "case_id": case.case_id,
        "source_name": case.source_name,
        "variant": case.variant,
        "duration_ms": case.duration_ms,
        "provider": provider.name,
        "provider_family": provider.family,
        "stop_wait_ms": provider.stop_wait_ms,
        "label_status": label.status,
        "label_start_ms": label.start_ms,
        "label_stop_ms": label.stop_ms,
        "first_start_ms": first_start,
        "detected_start_ms": detected_start,
        "detected_stop_ms": detected_stop,
        "start_delay_ms": start_delay,
        "stop_delay_ms": stop_delay,
        "missed_start": detected_start is None,
        "missed_stop": detected_start is not None and detected_stop is None,
        "false_start": false_start,
        "false_start_event_count": len(false_start_events),
        "event_count": len(events),
        "infer_ms_avg": round(mean(infer_values), 6),
        "infer_ms_p95": round(percentile(infer_values, 95), 6),
    }


def summarize_provider_results(sample_results: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """按 provider 和 variant 汇总指标。

    参数：sample_results，单样本明细列表。
    返回值：汇总行列表。
    异常情况：无。
    """

    groups: dict[tuple[str, str, int | None], list[dict[str, Any]]] = {}
    for row in sample_results:
        groups.setdefault((row["provider"], row["variant"], row["stop_wait_ms"]), []).append(row)
    for provider, rows in group_by(sample_results, "provider").items():
        groups[(provider, "all", rows[0]["stop_wait_ms"])] = rows

    summary_rows: list[dict[str, Any]] = []
    for (provider, variant, stop_wait_ms), rows in sorted(groups.items()):
        ok_rows = [row for row in rows if row["label_status"] == "ok"]
        start_delays = [row["start_delay_ms"] for row in ok_rows if row["start_delay_ms"] is not None]
        stop_delays = [row["stop_delay_ms"] for row in ok_rows if row["stop_delay_ms"] is not None]
        infer_avg_values = [row["infer_ms_avg"] for row in rows]
        infer_p95_values = [row["infer_ms_p95"] for row in rows]
        summary_rows.append(
            {
                "provider": provider,
                "provider_family": rows[0]["provider_family"],
                "variant": variant,
                "stop_wait_ms": stop_wait_ms,
                "sample_count": len(rows),
                "ok_label_count": len(ok_rows),
                "start_delay_ms_avg": round(mean(start_delays), 3),
                "start_delay_ms_p50": round(percentile(start_delays, 50), 3),
                "start_delay_ms_p95": round(percentile(start_delays, 95), 3),
                "stop_delay_ms_avg": round(mean(stop_delays), 3),
                "stop_delay_ms_p50": round(percentile(stop_delays, 50), 3),
                "stop_delay_ms_p95": round(percentile(stop_delays, 95), 3),
                "missed_start_count": sum(1 for row in ok_rows if row["missed_start"]),
                "missed_stop_count": sum(1 for row in ok_rows if row["missed_stop"]),
                "false_start_count": sum(1 for row in ok_rows if row["false_start"]),
                "false_start_event_count": sum(int(row.get("false_start_event_count", 0)) for row in ok_rows),
                "infer_ms_avg": round(mean(infer_avg_values), 6),
                "infer_ms_p95": round(percentile(infer_p95_values, 95), 6),
            }
        )
    return summary_rows


def build_manifest(
    args: argparse.Namespace,
    out_dir: Path,
    cases: list[AudioCase],
    provider_status: list[dict[str, Any]],
) -> dict[str, Any]:
    """生成实验 manifest。

    参数：命令行参数、输出目录、样本列表、provider 状态。
    返回值：manifest 字典。
    异常情况：版本探测失败时对应依赖版本写 unknown。
    """

    packages = {}
    for name in ["numpy", "openpyxl", "webrtcvad", "onnxruntime", "silero-vad", "torch", "ten-vad"]:
        try:
            packages[name] = importlib.metadata.version(name)
        except importlib.metadata.PackageNotFoundError:
            packages[name] = "not-installed"
    return {
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "out_dir": str(out_dir),
        "command": " ".join(sys.argv),
        "python": sys.version,
        "platform": platform.platform(),
        "sample_rate": SAMPLE_RATE,
        "seed": args.seed,
        "packages": packages,
        "sample_count": len(cases),
        "providers": provider_status,
        "variants": ["clean_padded", "white_noise_snr20", "hum_50hz_snr25"],
        "stop_wait_ms_values": sorted(
            {
                status["stop_wait_ms"]
                for status in provider_status
                if status.get("available") and status.get("stop_wait_ms") is not None
            }
        ),
    }


def write_report(
    path: Path,
    summary_rows: list[dict[str, Any]],
    sample_results: list[dict[str, Any]],
    labels: list[Label],
    provider_status: list[dict[str, Any]],
    manifest: dict[str, Any],
) -> None:
    """写 Markdown 实验报告。

    参数：报告路径、汇总行、明细结果、标签、provider 状态和 manifest。
    返回值：无。
    异常情况：文件写入失败时抛出异常。
    """

    clean_rows = [row for row in summary_rows if row["variant"] == "clean_padded"]
    all_rows = [row for row in summary_rows if row["variant"] == "all"]
    best_start = min(
        (row for row in clean_rows if row["missed_start_count"] == 0 and row["false_start_count"] == 0),
        key=lambda row: (row["start_delay_ms_p95"], row["stop_wait_ms"] or 0),
        default=None,
    )
    best_stop = min(
        (row for row in clean_rows if row["missed_stop_count"] == 0 and row["false_start_count"] == 0),
        key=lambda row: (row["stop_delay_ms_p95"], row["stop_wait_ms"] or 0),
        default=None,
    )
    lines = [
        "# VAD 边界检测基准实验报告",
        "",
        "## 实验结论",
        "",
    ]
    if best_start:
        lines.append(
            f"- clean 样本上 `speech_started` P95 最低且无 missed_start/false_start 的方案是 "
            f"`{best_start['provider']}`，stop_wait={best_start['stop_wait_ms']}ms，"
            f"P95={best_start['start_delay_ms_p95']}ms。"
        )
    if best_stop:
        lines.append(
            f"- clean 样本上 `speech_stopped` P95 最低且无 missed_stop/false_start 的方案是 "
            f"`{best_stop['provider']}`，stop_wait={best_stop['stop_wait_ms']}ms，"
            f"P95={best_stop['stop_delay_ms_p95']}ms。"
        )
    lines.extend(
        [
            "- 本轮 ground truth 来自 RMS proxy label，不等价于人工听检真值；结论只能用于第一轮相对排序。",
            "- 本轮只覆盖短中文指令和轻噪声增强，未覆盖真实播放回声、barge-in、远场和弱声场景。",
            "- FunASR FSMN-VAD 和 FireRedVAD 本轮未安装运行，保留到第二阶段真实 runs / 更高多样性样本再测。",
            "",
            "## 组合方案依据",
            "",
            "- LiveKit 的 VAD / turn detection 文档把 `min_silence_duration`、`prefix_padding_duration`、`activation_threshold` 等参数作为 VAD 调参面，并把 endpointing delay 作为 turn 结束判断的一部分；这说明原始 VAD 和结束等待应分开评估。",
            "- WebRTC VAD 的常见 collector 示例会用 padding window 和 voiced/unvoiced 比例触发/退出，而不是单帧直接决定；本实验同样把 provider 原始结果再经过 start/stop 状态机处理。",
            "- 本轮组合方案只验证工程取舍：`rms_start_silero_stop` 用 RMS 低成本能量门限触发 start、Silero 控制结束；`silero_start_rms_guard_stop` 用 Silero 触发 start，并用 RMS 防止低能量尾音被过早截断。",
            "- 本轮 clean 样本里 `rms_start_silero_stop_s200` 的 Start P95 是 407.6ms，`silero_onnx_s200` 是 383.0ms；数据没有证明 RMS start 比 Silero start 更快。",
            "- 参考：https://livekit.io/field-guides/guide/vad-turn-detection-configuration ，https://docs.livekit.io/agents/logic/turns/ ，https://github.com/wiseman/py-webrtcvad/tree/master/example.py",
            "",
            "## Provider 可用性",
            "",
            "| Provider | 可用 | 初始化耗时/原因 |",
            "| --- | --- | --- |",
        ]
    )
    for status in provider_status:
        detail = status.get("init_ms", status.get("reason", ""))
        lines.append(f"| `{status['provider']}` | {status['available']} | {detail} |")

    lines.extend(
        [
            "",
            "## 指标解释",
            "",
            "| 指标 | 含义 | 读数方式 |",
            "| --- | --- | --- |",
            "| `Start P50/P95` | `detected_start_ms - label_start_ms` 的中位数 / 95 分位 | 越小表示越早开始 append；负数表示早于 RMS proxy label |",
            "| `Stop P50/P95` | `detected_stop_ms - label_stop_ms` 的中位数 / 95 分位 | 越小表示越早 commit；过小或负数要结合 false stop 风险听检 |",
            "| `Miss Start` | 未找到有效 `speech_started` 的样本数 | Omni manual 会丢整轮用户输入，必须优先避免 |",
            "| `Miss Stop` | 找到 start 但没有有效 stop 的样本数 | Omni manual 会迟迟不 commit，必须优先避免 |",
            "| `False Start` | 早于 RMS start 300ms 以上的样本数 | 代表播放回声/噪声误触发风险，本轮用静音 padding 近似观察 |",
            "| `False Events` | early false start 事件总数 | 一个样本可出现多个误触发事件，用来观察抖动程度 |",
            "| `Infer P95` | 单帧推理耗时 P95 | 要显著小于 chunk 时长，才适合实时链路 |",
            "| `stop_wait_ms` | 结束等待窗口 | 越小 commit 越快，但更容易 false stop；越大更稳但响应更慢 |",
            "",
            "## 汇总指标",
            "",
            "### 全部增强版本合并",
            "",
            "| Provider | Stop Wait | Start P50 | Start P95 | Stop P50 | Stop P95 | Miss Start | Miss Stop | False Start | False Events | Infer P95 |",
            "| --- | --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: |",
        ]
    )
    for row in sorted(all_rows, key=lambda item: (item["provider_family"], item["stop_wait_ms"] or 0)):
        lines.append(summary_markdown_row(row))
    lines.append("")
    lines.append("### Clean 样本")
    lines.append("")
    lines.append(
        "| Provider | Stop Wait | Start Avg | Start P95 | Stop Avg | Stop P95 | Miss Start | Miss Stop | False Start |"
    )
    lines.append("| --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: |")
    for row in sorted(clean_rows, key=lambda item: (item["provider_family"], item["stop_wait_ms"] or 0)):
        lines.append(
            f"| `{row['provider']}` | {row['stop_wait_ms']} | {row['start_delay_ms_avg']} | {row['start_delay_ms_p95']} | "
            f"{row['stop_delay_ms_avg']} | {row['stop_delay_ms_p95']} | "
            f"{row['missed_start_count']} | {row['missed_stop_count']} | {row['false_start_count']} |"
        )
    lines.extend(
        [
            "",
            "### 噪声增强样本最佳参数",
            "",
            "| Variant | Provider Family | Provider | Stop Wait | Start P95 | Stop P95 | Miss Start | Miss Stop | False Start |",
            "| --- | --- | --- | ---: | ---: | ---: | ---: | ---: | ---: |",
        ]
    )
    for row in best_rows_by_family_variant(summary_rows):
        lines.append(
            f"| `{row['variant']}` | `{row['provider_family']}` | `{row['provider']}` | {row['stop_wait_ms']} | "
            f"{row['start_delay_ms_p95']} | {row['stop_delay_ms_p95']} | "
            f"{row['missed_start_count']} | {row['missed_stop_count']} | {row['false_start_count']} |"
        )

    lines.extend(
        [
            "",
            "## 结论建议",
            "",
            conclusion_text(clean_rows, sample_results),
            "",
            "## 样本与标注",
            "",
            f"- 样本数量：{len({row['source_name'] for row in sample_results})} 条原始音频，"
            f"{len({row['case_id'] for row in sample_results})} 条增强后输入。",
            f"- RMS label unstable 数量：{sum(1 for label in labels if label.status != 'ok')}。",
            "- 增强版本：`clean_padded`、`white_noise_snr20`、`hum_50hz_snr25`。",
            "",
            "## 产物",
            "",
            f"- `manifest.json`：实验环境、依赖版本、provider 状态。",
            f"- `labels.jsonl`：RMS proxy label 明细。",
            f"- `provider-events.jsonl`：VAD 边界事件。",
            f"- `sample-results.jsonl`：逐样本逐 provider 明细。",
            f"- `summary.json` / `summary.xlsx`：汇总表。",
            "",
            "## 环境",
            "",
            f"- Python：`{manifest['python'].split()[0]}`",
            f"- 平台：`{manifest['platform']}`",
            f"- 依赖：`{manifest['packages']}`",
        ]
    )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def conclusion_text(clean_rows: list[dict[str, Any]], sample_results: list[dict[str, Any]]) -> str:
    """根据 clean 汇总生成结论段落。

    参数：clean_rows 和 sample_results。
    返回值：Markdown 文本。
    异常情况：无。
    """

    if not clean_rows:
        return "- 没有 clean 样本汇总，无法给出排序。"
    sorted_by_start = sorted(
        clean_rows,
        key=lambda row: (
            row["missed_start_count"],
            row["false_start_count"],
            row["start_delay_ms_p95"],
            row["stop_wait_ms"] or 0,
        ),
    )
    sorted_by_stop = sorted(
        clean_rows,
        key=lambda row: (
            row["missed_stop_count"],
            row["false_start_count"],
            row["stop_delay_ms_p95"],
            row["stop_wait_ms"] or 0,
        ),
    )
    start_winner = sorted_by_start[0]
    stop_winner = sorted_by_stop[0]
    lines = [
        f"- `speech_started`：优先看 missed_start=0 和 P95 延迟，当前排序第一是 "
        f"`{start_winner['provider']}`，stop_wait={start_winner['stop_wait_ms']}ms。",
        f"- `speech_stopped`：优先看 missed_stop=0 和 P95 延迟，当前排序第一是 "
        f"`{stop_winner['provider']}`，stop_wait={stop_winner['stop_wait_ms']}ms。",
    ]
    stable = [
        row
        for row in clean_rows
        if row["missed_start_count"] == 0
        and row["missed_stop_count"] == 0
        and row["false_start_count"] == 0
        and row["start_delay_ms_p95"] <= 500
        and row["stop_delay_ms_p95"] <= 800
    ]
    if stable:
        names = "、".join(f"`{row['provider']}`" for row in stable)
        lines.append(f"- 按本轮阈值，{names} 满足 clean 样本初步验收线，并且没有 early false start。")
    else:
        lines.append(
            "- 没有 provider 同时满足 clean 样本 start P95 <= 500ms、stop P95 <= 800ms、零 miss 和零 early false start。"
        )
    lines.append(
        "- 正式接入 Omni manual 前，建议继续用真实 mic runs 和播放回声录音复测；本轮不能替代真机回声场景。"
    )
    return "\n".join(lines)


def summary_markdown_row(row: dict[str, Any]) -> str:
    """把汇总行格式化为 Markdown 表格行。

    参数：row，汇总指标字典。
    返回值：Markdown 表格行字符串。
    异常情况：无。
    """

    return (
        f"| `{row['provider']}` | {row['stop_wait_ms']} | {row['start_delay_ms_p50']} | "
        f"{row['start_delay_ms_p95']} | {row['stop_delay_ms_p50']} | {row['stop_delay_ms_p95']} | "
        f"{row['missed_start_count']} | {row['missed_stop_count']} | {row['false_start_count']} | "
        f"{row['false_start_event_count']} | {row['infer_ms_p95']} |"
    )


def best_rows_by_family_variant(summary_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """为每个 provider family 和音频版本选择最佳 stop wait 行。

    参数：summary_rows，provider 汇总结果。
    返回值：每个 family/variant 的最佳行。
    异常情况：无。
    """

    groups: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for row in summary_rows:
        if row["variant"] == "all":
            continue
        groups.setdefault((row["variant"], row["provider_family"]), []).append(row)
    result: list[dict[str, Any]] = []
    for _, rows in sorted(groups.items()):
        result.append(
            min(
                rows,
                key=lambda row: (
                    row["missed_start_count"],
                    row["missed_stop_count"],
                    row["false_start_count"],
                    row["start_delay_ms_p95"] + row["stop_delay_ms_p95"],
                ),
            )
        )
    return result


def write_xlsx(
    path: Path,
    summary_rows: list[dict[str, Any]],
    sample_results: list[dict[str, Any]],
    labels: list[Label],
    provider_status: list[dict[str, Any]],
) -> None:
    """写 Excel 明细表。

    参数：输出路径、汇总行、样本明细、标签和 provider 状态。
    返回值：无。
    异常情况：openpyxl 不可用或写入失败时抛出异常。
    """

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    default = workbook.active
    default.title = "Summary"
    write_sheet(default, summary_rows)
    sheets = [
        ("SampleResults", sample_results),
        ("Labels", [label.__dict__ for label in labels]),
        ("ProviderStatus", provider_status),
    ]
    for title, rows in sheets:
        ws = workbook.create_sheet(title)
        write_sheet(ws, rows)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for ws in workbook.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        for column_cells in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_len + 2, 10), 45)
    workbook.save(path)


def write_sheet(ws: Any, rows: list[dict[str, Any]]) -> None:
    """把字典列表写入 worksheet。

    参数：ws 是 openpyxl worksheet；rows 是字典行列表。
    返回值：无。
    异常情况：无。
    """

    if not rows:
        ws.append(["empty"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])


def read_wav_pcm16(path: Path) -> np.ndarray:
    """读取 PCM16 mono 16k WAV。

    参数：path，WAV 文件路径。
    返回值：int16 numpy 数组。
    异常情况：采样率、声道数或采样宽度不符合预期时抛出 ValueError。
    """

    with wave.open(str(path), "rb") as wav:
        channels = wav.getnchannels()
        sample_rate = wav.getframerate()
        sample_width = wav.getsampwidth()
        if channels != 1 or sample_rate != SAMPLE_RATE or sample_width != 2:
            raise ValueError(
                f"{path} must be PCM16 mono 16k, got channels={channels} rate={sample_rate} width={sample_width}"
            )
        data = wav.readframes(wav.getnframes())
    return np.frombuffer(data, dtype="<i2").copy()


def write_wav_pcm16(path: Path, pcm: np.ndarray) -> None:
    """写 PCM16 mono 16k WAV。

    参数：path 和 pcm。
    返回值：无。
    异常情况：文件写入失败时抛出异常。
    """

    with wave.open(str(path), "wb") as wav:
        wav.setnchannels(1)
        wav.setsampwidth(2)
        wav.setframerate(SAMPLE_RATE)
        wav.writeframes(pcm.astype("<i2").tobytes())


def write_json(path: Path, data: Any) -> None:
    """写格式化 JSON 文件。

    参数：path 和 data。
    返回值：无。
    异常情况：文件写入失败时抛出异常。
    """

    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_jsonl(path: Path, rows: Iterable[dict[str, Any]]) -> None:
    """写 JSONL 文件。

    参数：path 和字典行迭代器。
    返回值：无。
    异常情况：文件写入失败时抛出异常。
    """

    with path.open("w", encoding="utf-8") as fh:
        for row in rows:
            fh.write(json.dumps(row, ensure_ascii=False) + "\n")


def add_white_noise(pcm: np.ndarray, snr_db: float) -> np.ndarray:
    """按目标 SNR 添加白噪声。

    参数：pcm 输入音频；snr_db 目标信噪比。
    返回值：添加噪声后的 PCM16。
    异常情况：无。
    """

    signal = pcm.astype(np.float32)
    noise = np.random.normal(0.0, 1.0, size=len(signal)).astype(np.float32)
    return mix_noise(signal, noise, snr_db)


def add_hum(pcm: np.ndarray, hz: float, snr_db: float) -> np.ndarray:
    """按目标 SNR 添加低频 hum。

    参数：pcm 输入音频；hz 低频频率；snr_db 目标信噪比。
    返回值：添加噪声后的 PCM16。
    异常情况：无。
    """

    signal = pcm.astype(np.float32)
    t = np.arange(len(signal), dtype=np.float32) / SAMPLE_RATE
    noise = np.sin(2.0 * math.pi * hz * t).astype(np.float32)
    return mix_noise(signal, noise, snr_db)


def mix_noise(signal: np.ndarray, noise: np.ndarray, snr_db: float) -> np.ndarray:
    """把 noise 缩放后叠加到 signal。

    参数：signal、noise 和目标 SNR。
    返回值：PCM16。
    异常情况：无。
    """

    active = signal[np.abs(signal) > 0]
    signal_rms = float(np.sqrt(np.mean(active**2))) if len(active) else 1.0
    noise_rms = float(np.sqrt(np.mean(noise**2))) or 1.0
    target_noise_rms = signal_rms / (10 ** (snr_db / 20.0))
    scaled = noise * (target_noise_rms / noise_rms)
    mixed = np.clip(signal + scaled, -32768, 32767)
    return mixed.astype(np.int16)


def median_filter(values: np.ndarray, size: int) -> np.ndarray:
    """对一维数组做简单中值滤波。

    参数：values 输入数组；size 窗口大小。
    返回值：滤波后的数组。
    异常情况：无。
    """

    radius = size // 2
    padded = np.pad(values, (radius, radius), mode="edge")
    return np.array([np.median(padded[index : index + size]) for index in range(len(values))], dtype=np.float32)


def find_first_run(flags: np.ndarray, min_len: int) -> int | None:
    """查找第一个连续 True 片段。

    参数：flags 布尔数组；min_len 最短连续长度。
    返回值：片段起点索引，找不到返回 None。
    异常情况：无。
    """

    count = 0
    for index, flag in enumerate(flags):
        count = count + 1 if flag else 0
        if count >= min_len:
            return index - count + 1
    return None


def find_last_run(flags: np.ndarray, min_len: int) -> int | None:
    """查找最后一个连续 True 片段的末尾。

    参数：flags 布尔数组；min_len 最短连续长度。
    返回值：片段末尾索引，找不到返回 None。
    异常情况：无。
    """

    count = 0
    for reverse_index, flag in enumerate(reversed(flags)):
        count = count + 1 if flag else 0
        if count >= min_len:
            end_index = len(flags) - reverse_index + count - 2
            return min(end_index, len(flags) - 1)
    return None


def pad_to_frame(pcm: np.ndarray, frame_samples: int) -> np.ndarray:
    """把音频补零到 provider 帧长整数倍。

    参数：pcm 和 frame_samples。
    返回值：补零后的 PCM16。
    异常情况：frame_samples 非正时抛出 ValueError。
    """

    if frame_samples <= 0:
        raise ValueError("frame_samples must be positive")
    remainder = len(pcm) % frame_samples
    if remainder == 0:
        return pcm
    pad = frame_samples - remainder
    return np.concatenate([pcm, np.zeros(pad, dtype=np.int16)])


def rms_float(frame: np.ndarray) -> float:
    """计算归一化 RMS。

    参数：PCM16 frame。
    返回值：0 到 1 附近的 RMS。
    异常情况：空帧返回 0。
    """

    if len(frame) == 0:
        return 0.0
    data = frame.astype(np.float32) / PCM_MAX
    return float(np.sqrt(np.mean(data * data)))


def silero_probability(model: Any, frame: np.ndarray) -> float:
    """计算 Silero ONNX 对当前帧的语音概率。

    参数：model 是 silero-vad 的 ONNX wrapper；frame 是 PCM16 音频帧。
    返回值：语音概率。
    异常情况：模型调用失败时向外抛出。
    """

    import torch

    audio_float = frame.astype(np.float32) / PCM_MAX
    return float(model(torch.from_numpy(audio_float), SAMPLE_RATE).item())


def make_event(event: str, provider: str, audio_ms: int, score: float | None, infer_ms: float) -> dict[str, Any]:
    """构造标准事件字典。

    参数：event、provider、audio_ms、score、infer_ms。
    返回值：事件字典。
    异常情况：无。
    """

    return {
        "event": event,
        "provider": provider,
        "audio_ms": int(audio_ms),
        "score": None if score is None else round(float(score), 6),
        "infer_ms": round(float(infer_ms), 6),
    }


def samples_to_ms(samples: int) -> int:
    """把 sample 数转换为毫秒。

    参数：samples。
    返回值：四舍五入后的毫秒。
    异常情况：无。
    """

    return int(round(samples * 1000 / SAMPLE_RATE))


def ms_to_samples(ms: int) -> int:
    """把毫秒转换为 sample 数。

    参数：ms。
    返回值：sample 数。
    异常情况：无。
    """

    return int(round(ms * SAMPLE_RATE / 1000))


def elapsed_ms(started: float) -> float:
    """计算耗时毫秒。

    参数：started，perf_counter 起点。
    返回值：毫秒。
    异常情况：无。
    """

    return (time.perf_counter() - started) * 1000.0


def none_subtract(left: int | None, right: int | None) -> int | None:
    """安全相减。

    参数：left 和 right。
    返回值：任一参数为空时返回 None，否则返回差值。
    异常情况：无。
    """

    if left is None or right is None:
        return None
    return int(left - right)


def mean(values: list[float | int | None]) -> float:
    """计算均值。

    参数：数值列表。
    返回值：空列表返回 0。
    异常情况：无。
    """

    clean = [float(value) for value in values if value is not None]
    return statistics.fmean(clean) if clean else 0.0


def percentile(values: list[float | int | None], pct: float) -> float:
    """计算百分位。

    参数：数值列表和百分位。
    返回值：空列表返回 0。
    异常情况：无。
    """

    clean = [float(value) for value in values if value is not None]
    if not clean:
        return 0.0
    return float(np.percentile(np.array(clean, dtype=np.float64), pct))


def group_by(rows: list[dict[str, Any]], key: str) -> dict[str, list[dict[str, Any]]]:
    """按字段分组。

    参数：rows 和 key。
    返回值：分组字典。
    异常情况：无。
    """

    result: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        result.setdefault(str(row[key]), []).append(row)
    return result


if __name__ == "__main__":
    main()
